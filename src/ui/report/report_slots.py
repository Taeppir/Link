import os
import platform
import subprocess
from datetime import datetime, timedelta
from PySide6.QtCore import QObject
from PySide6.QtWidgets import QTableWidgetItem, QFileDialog, QMessageBox
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

class ReportSlots(QObject):
    def __init__(self,panel):
        super().__init__()
        self.panel = panel

    def update_report(self, data:dict):
        print("[ReportPanel] Report 업데이트 시작")

        self.panel.showReport(True)

        distance_km = data.get("distance", 0)
        time_h = data.get("time", 0)
        fuel_ton = data.get("fuel", 0)
        speed_mps = 8.0

        optimal_distance_km = data.get("optimal_distance", distance_km)
        optimal_time_h = data.get("optimal_time", time_h)
        optimal_fuel_ton = data.get("optimal_fuel", fuel_ton)

        start_time = datetime.now()
        end_time = start_time + timedelta(hours=time_h)
        start_str = start_time.strftime("%Y-%m-%d %H:%M")
        end_str = end_time.strftime("%Y-%m-%d %H:%M")

        result = {
            "compare": [
                # [출발, 도착, 거리, 시간, 연료, 속력]
                [start_str, "", f"{optimal_distance_km:.2f}", f"{optimal_time_h:.2f}", f"{optimal_fuel_ton:.2f}", f"{speed_mps:.1f}"],
                [start_str, end_str, f"{distance_km:.2f}", f"{time_h:.2f}", f"{fuel_ton:.2f}", f"{speed_mps:.1f}"]
            ],
            "optimal": {
                "summary": [start_str, "", f"{optimal_distance_km:.2f}", f"{optimal_time_h:.2f}", f"{optimal_fuel_ton:.2f}", f"{speed_mps:.1f}"],
                "fuel": optimal_fuel_ton
            },
            "shortest": {
                "summary": [start_str, end_str, f"{distance_km:.2f}", f"{time_h:.2f}", f"{fuel_ton:.2f}", f"{speed_mps:.1f}"],
                "fuel": fuel_ton
            }
        }

        # 비교 탭
        compare_table = self.panel.compare_tab.summary
        for row, data_row in enumerate(result["compare"]):
            for column, value in enumerate(data_row):
                compare_table.setItem(row, column, QTableWidgetItem(str(value)))

        # 연료 절감 문구
        optimal_fuel = result["optimal"]["fuel"]
        shortest_fuel = result["shortest"]["fuel"]
        fuel_saving = shortest_fuel - optimal_fuel
        fuel_percent = (fuel_saving / shortest_fuel) * 100 if shortest_fuel != 0 else 0

        fuel_saving = shortest_fuel - optimal_fuel
        fuel_percent = (fuel_saving / shortest_fuel) * 100

        summary_text = self.panel.compare_tab.summary_text
        summary_text.setText(
            f"최적 경로 선택 시 최단 경로 대비 "
            f"<span style='color:#d9534f; font-weight:bold;'>🔻 약 {fuel_percent:.1f}%</span> "
            f"({fuel_saving:.1f}톤) 연료 절감이 예상됩니다."
        )

        # 최적 탭
        optimal_summary = self.panel.optimal_tab.summary
        for column, value in enumerate(result["optimal"]["summary"]):
            optimal_summary.setItem(0, column, QTableWidgetItem(str(value)))

        # 최단 탭
        shortest_summary = self.panel.shortest_tab.summary
        for column, value in enumerate(result["shortest"]["summary"]):
            shortest_summary.setItem(0, column, QTableWidgetItem(str(value)))

        path_points = data.get("path_points", [])
        n = len(path_points)

        if n <= 100:
            sampled_points = path_points
        else:
            step = n / 100 
            sampled_points = [path_points[int(i * step)] for i in range(100)]

        shortest_detail = self.panel.shortest_tab.detail
        if shortest_detail is not None:
            shortest_detail.clearContents()
            shortest_detail.setRowCount(len(sampled_points))

        for i, point in enumerate(sampled_points):
            lat = f"{point['lat']:.6f}"
            lon = f"{point['lon']:.6f}"

            shortest_detail.setItem(i, 2, QTableWidgetItem(lat))
            shortest_detail.setItem(i, 3, QTableWidgetItem(lon))
            shortest_detail.setItem(i, 4, QTableWidgetItem(""))
            shortest_detail.setItem(i, 5, QTableWidgetItem(""))
            shortest_detail.setItem(i, 6, QTableWidgetItem(f"{speed_mps:.1f}"))


        print("[ReportPanel] Report 업데이트 완료")

    def download_report(self):
        # -- 파일 생성 --
        file_path, _ = QFileDialog.getSaveFileName(
            self.panel, "탐색 결과 다운로드", "report.xlsx", "Excel Files (*.xlsx)")
        
        if not file_path:
            return
        
        wb = Workbook()
        wb.remove(wb.active)

        # -- 시트 생성 --
        compare_tab = self.panel.compare_tab
        self._write_table_to_sheet(wb, "비교", compare_tab.summary)

        optimal_tab = self.panel.optimal_tab
        self._write_table_to_sheet(wb,"최적경로_상세", optimal_tab.detail)

        shortest_tab = self.panel.shortest_tab
        self._write_table_to_sheet(wb,"최단경로_상세", shortest_tab.detail)
        
        wb.save(file_path)
        print(f"[ReportPanel] Excel 저장 완료: {file_path}")

        answer = QMessageBox.question(
            self.panel,
            "파일 저장 완료",
            f"결과 파일이 저장되었습니다.\n\n{file_path}\n\n지금 파일을 열겠습니까?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )

        if answer != QMessageBox.Yes:
            return
        
        try:
            system = platform.system()
            if system == "Windows":
                os.startfile(file_path)
            elif system == "Darwin":  # macOS
                subprocess.run(["open", file_path])
            else:  # Linux
                subprocess.run(["xdg-open", file_path])
            print("[ReportPanel] Excel 자동 실행 완료")
        except Exception as e:
            QMessageBox.warning(self.panel, "오류", f"파일을 열 수 없습니다.\n\n{e}")
            print(f"[ReportPanel] Excel 자동 실행 실패: {e}")

    def _write_table_to_sheet(self, wb, sheet_name, table):
        ws = wb.create_sheet(sheet_name)

        col_count = table.columnCount()
        row_count = table.rowCount()

        # -- 헤더 --
        headers = [""]
        for column in range(col_count):
            header_item = table.horizontalHeaderItem(column)
            headers.append(header_item.text() if header_item else "")
        ws.append(headers)

        # -- 데이터 --
        for row in range(row_count):
            datas = []

            row_header = table.verticalHeaderItem(row)
            datas.append(row_header.text() if row_header else "")

            for column in range(col_count):
                item = table.item(row, column)
                datas.append(item.text() if item else "")
            ws.append(datas)

        # -- 정렬 --
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        min_width = 30
        for col_idx in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_cell = ws.cell(row=1, column=col_idx)

            if header_cell.value:
                text = str(header_cell.value)
                width = len(text) * 2

                if 2 <= col_idx <= 3:
                    ws.column_dimensions[col_letter].width = max(min_width, width)
                else:
                    ws.column_dimensions[col_letter].width = width